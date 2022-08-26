from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from json import dumps as encodeJson, loads as decodeJson
from uuid import UUID
from functools import reduce
import re
from datetime import datetime
import sys
import os


def metaformat(m):
    return m if 'metadata.' in m else f'metadata.{m}'


def disablePrint():
    sys.stdout = open(os.devnull, 'w')


def enablePrint():
    sys.stdout = sys.__stdout__


def readJSON(file):
    with open(file, 'r') as file:
        ret = decodeJson(file.read())
    return ret


def unfold(obj, ret={}, prefix=''):
    if type(obj) == dict:
        for x in obj:
            if type(obj[x]) != dict or obj[x] == {}:
                ret[f'{prefix}{"." if prefix != "" else ""}{x}'] = obj[x]
            else:
                ret = unfold(obj[x], ret, f'{prefix}{"." if prefix != "" else ""}{x}')
    return ret


def fold(obj):
    ret, c = {}, "'"
    for entry in obj:
        tmp = entry.split('.')
        for x in range(len(tmp)):
            if not eval(f'\'{tmp[x]}\' in ret{"".join(f"[{c}{k}{c}]" for k in tmp[:x])}'):
                exec(f'ret{"".join(f"[{c}{k}{c}]" for k in tmp[:x+1])} = {{}}')
        exec(f'ret{"".join(f"[{c}{k}{c}]" for k in tmp)} = obj[entry]')
    return ret


def getAllKeys(arr):
    keys = set()
    for item in arr:
        for key in item.keys():
            try:
                keys.add(key)
            except Exception:
                pass
    return list(keys)


def qs(var):
    return f"'{var}'" if type(var) == str else str(var)


def xlsxFormatter(sheet):
    fr = sheet['B2']
    sheet.freeze_panes = fr
    sheet.auto_filter.ref = sheet.dimensions
    custom_header = NamedStyle(name="custom_header")
    custom_header.fill = PatternFill("solid", fgColor="000000")
    custom_header.font = Font(color="FFFFFF")
    custom_header.alignment = Alignment(vertical="center")
    normal_cell = NamedStyle(name="normal_cell")
    normal_cell.alignment = Alignment(vertical="center")
    firstRow = sheet['1']
    for cell in firstRow:
        cell.style = custom_header
    sheet.row_dimensions[1].height = 23
    for i, row in enumerate([c for c in sheet.iter_rows()][1:]):
        sheet.row_dimensions[i+2].height = 23
        for cell in row:
            cell.style = normal_cell
    for coll in sheet.columns:
        length = max(len(str(cell.value).split('\n')[0]) for cell in coll)
        sheet.column_dimensions[coll[0].column_letter].width = length if length > 15 else 15


def dateFromISO(date):
    if date is not None and type(date) == str:
        try:
            if date[-1] == 'Z':
                date = date[:-1]+'+0000'
            else:
                z = date[-6:]
                date = date[:-6] + z.replace(':', '')
            print(date)
            return datetime.strptime(date, '%Y-%m-%dT%H:%M:%S.%f%z')
        except Exception as e:
            print(e)
            return datetime.now()
    else:
        return datetime.now()


def matchRgx(txt, pattern):
    pattern = pattern if type(pattern) == re.Pattern else re.compile(pattern)
    ret = []
    for match in pattern.finditer(txt):
        st, ed = match.span()
        ret.append(match.string[st:ed])
    return ret


def uncomment(line):
    li = line
    rx = re.compile(r'^[\s]{0,3}#[\s]{0,3}|#[\s]{0,3}\b|#')
    matches = matchRgx(line, rx)
    if len(matches) > 0:
        li = li.replace(matches[0], '', 1)
    return li


def deleteComment(txt):
    ret = txt
    rx = re.compile(r'[\t|\s]+#[^\n]+$')
    for m in matchRgx(txt, rx):
        ret = ret.replace(m, '')
    return ret
